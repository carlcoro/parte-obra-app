import streamlit as st
import openpyxl
import pandas as pd
from fpdf import FPDF
import io
import datetime

# ==========================================
# 1. CONFIGURACIÓN E INTERFAZ DE STREAMLIT
# ==========================================
st.set_page_config(page_title="Gestor de Partes de Obra", layout="wide")
st.title("🚧 Sistema de Gestión de Partes de Obra")

# Diccionario global de unificación de nombres (¡Aquí está definido de forma segura!)
DICCIONARIO_NOMBRES = {
    "carlos": "Carlos Corobo", "carlos c": "Carlos Corobo", "carlos corobo": "Carlos Corobo",
    "sergio": "Sergio Moreno", "sergio m": "Sergio Moreno", "sergio moreno": "Sergio Moreno",
    "arturo": "Arturo Cano", "arturo c": "Arturo Cano", "arturo cano": "Arturo Cano",
    "gabriel": "Gabriel Moreno", "gabriel m": "Gabriel Moreno", "gabriel moreno": "Gabriel Moreno"
}

def unificar_nombre(nombre_crudo):
    if not nombre_crudo:
        return "Desconocido"
    return DICCIONARIO_NOMBRES.get(str(nombre_crudo).strip().lower(), str(nombre_crudo).strip())

def limpiar_texto_pdf(texto):
    """Reemplaza caracteres conflictivos para evitar errores en fuentes estándar de FPDF"""
    if not texto:
        return ""
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'ñ': 'n', 'Ñ': 'N', 'ü': 'u', 'Ü': 'U'
    }
    t = str(texto)
    for orig, dest in reemplazos.items():
        t = t.replace(orig, dest)
    return t

# ==========================================
# 2. FORMULARIO DE ENTRADA DE DATOS (WEB)
# ==========================================
col1, col2, col3 = st.columns(3)
with col1:
    num_parte = st.text_input("Nº Parte", value="1001")
    obra = st.text_input("Obra", value="Sin Obra")
    cliente = st.text_input("Cliente", value="")
with col2:
    fecha_dt = st.date_input("Fecha", datetime.date.today())
    fecha = fecha_dt.strftime("%d-%m-%Y")
    jefe_obra = st.text_input("Jefe de Obra", value="")
with col3:
    ubicacion = st.text_input("Ubicación", value="")
    km = st.number_input("Kilómetros realizados", min_value=0, value=0)

trabajos = st.text_area("Trabajos Realizados (Introduce los detalles de lo que se ha hecho)", value="")

st.subheader("👥 Personal Asignado")
df_operarios = st.data_editor(
    pd.DataFrame([{"Operario": "", "Horas": 0.0}]),
    num_rows="dynamic",
    key="tabla_operarios"
)

st.subheader("📦 Materiales Utilizados")
df_materiales = st.data_editor(
    pd.DataFrame([{"Material": "", "Cantidad": 0.0, "Unidad": "uds"}]),
    num_rows="dynamic",
    key="tabla_materiales"
)

st.subheader("📁 Archivo Base")
archivo_subido = st.file_uploader("Sube tu archivo 'Plantilla_Parte_Obra_Profesional.xlsx'", type=["xlsx"])

# ==========================================
# 3. LÓGICA DE PROCESAMIENTO AL PULSAR BOTÓN
# ==========================================
if archivo_subido is not None:
    if st.button("🚀 Generar Documentos y Actualizar Base de Datos"):
        
        # --- PROCESAR OPERARIOS Y MATERIALES ---
        operarios_datos_crudos = []
        operarios_lista = []
        for index, row in df_operarios.iterrows():
            if row["Operario"]:
                op_oficial = unificar_nombre(row["Operario"])
                operarios_datos_crudos.append((op_oficial, row["Horas"]))
                operarios_lista.append(f"{op_oficial} ({row['Horas']}h)")
        operarios_texto = ", ".join(operarios_lista)

        materiales_lista = []
        materiales_datos = []
        for index, row in df_materiales.iterrows():
            if row["Material"]:
                materiales_lista.append(f"{row['Material']}: {row['Cantidad']} {row['Unidad']}")
                materiales_datos.append((str(row['Material']), str(row['Cantidad']), str(row['Unidad'])))
        materiales_texto = ", ".join(materiales_lista)

        # --- GENERAR PDF ---
        class PDFParte(FPDF):
            def header(self):
                self.set_font("Helvetica", "B", 14)
                self.set_text_color(46, 125, 50)
                self.cell(190, 10, "PARTE DIARIO DE TRABAJO", border=1, ln=1, align="C")
                self.ln(5)

        pdf = PDFParte(orientation="P", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_font("Helvetica", "", 10)

        # Bloque Información General
        pdf.set_fill_color(232, 245, 233)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Nº Parte:", border=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(65, 7, limpiar_texto_pdf(num_parte), border=1)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Fecha:", border=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(65, 7, limpiar_texto_pdf(fecha), border=1, ln=1)

        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Obra:", border=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(65, 7, limpiar_texto_pdf(obra), border=1)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Jefe de Obra:", border=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(65, 7, limpiar_texto_pdf(jefe_obra), border=1, ln=1)

        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Ubicacion:", border=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(65, 7, limpiar_texto_pdf(ubicacion), border=1)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Cliente:", border=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(65, 7, limpiar_texto_pdf(cliente), border=1, ln=1)
        pdf.ln(5)

        # Trabajos
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(25, 110, 30)
        pdf.cell(190, 7, "TRABAJOS REALIZADOS", ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(190, 6, limpiar_texto_pdf(trabajos) if trabajos else "Ninguno", border=1)
        pdf.ln(5)

        # Operarios en PDF
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(25, 110, 30)
        pdf.cell(190, 7, "PERSONAL ASIGNADO", ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(130, 6, "Operario", border=1, fill=True)
        pdf.cell(60, 6, "Horas", border=1, ln=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        if not operarios_datos_crudos:
            pdf.cell(190, 6, "No se registro personal", border=1, ln=1)
        for op, hr in operarios_datos_crudos:
            pdf.cell(130, 6, limpiar_texto_pdf(op), border=1)
            pdf.cell(60, 6, f"{hr} h", border=1, ln=1)
        pdf.ln(5)

        # Materiales en PDF
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(25, 110, 30)
        pdf.cell(190, 7, "MATERIALES UTILIZADOS", ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(100, 6, "Material", border=1, fill=True)
        pdf.cell(45, 6, "Cantidad", border=1, fill=True)
        pdf.cell(45, 6, "Unidad", border=1, ln=1, fill=True)
        pdf.set_font("Helvetica", "", 10)
        if not materiales_datos:
            pdf.cell(190, 6, "No se registraron materiales", border=1, ln=1)
        for mat, cant, uni in materiales_datos:
            pdf.cell(100, 6, limpiar_texto_pdf(mat), border=1)
            pdf.cell(45, 6, limpiar_texto_pdf(cant), border=1)
            pdf.cell(45, 6, limpiar_texto_pdf(uni), border=1, ln=1)
        pdf.ln(6)
        
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(190, 6, f"Kilometros realizados: {km} km", ln=1)

        # Firmas
        pdf.ln(15)
        pos_y = pdf.get_y()
        pdf.line(20, pos_y + 15, 80, pos_y + 15)
        pdf.line(130, pos_y + 15, 190, pos_y + 15)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_y(pos_y + 16)
        pdf.cell(95, 5, "Firma del Cliente", align="C")
        pdf.cell(95, 5, "Firma del Jefe de Obra", align="C", ln=1)

        pdf_output = pdf.output(dest='S').encode('utf-8', errors='replace')

        # --- MODIFICAR EXCEL EN MEMORIA ---
        wb = openpyxl.load_workbook(archivo_subido)
        
        if "Base de Datos" not in wb.sheetnames:
            ws_base = wb.create_sheet(title="Base de Datos")
            ws_base.append(["Fecha", "Nº Parte", "Obra", "Cliente", "Ubicación", "Operarios", "Materiales", "Km", "Procesado"])
        else:
            ws_base = wb["Base de Datos"]
            
        nueva_fila = [fecha, num_parte, obra, cliente, ubicacion, operarios_texto, materiales_texto, km, "No"]
        ws_base.append(nueva_fila)

        if "Historial_Horas" not in wb.sheetnames:
            ws_horas = wb.create_sheet(title="Historial_Horas")
            ws_horas.append(["Fecha", "Nº Parte", "Obra", "Nombre Operario", "Horas Trabajadas"])
        else:
            ws_horas = wb["Historial_Horas"]

        for nombre_op, horas_op in operarios_datos_crudos:
            ws_horas.append([fecha, num_parte, obra, nombre_op, horas_op])

        # Limpieza de columnas G y H para recalcular totales de forma segura
        for fila in range(1, ws_horas.max_row + 1):
            ws_horas[f"G{fila}"] = None
            ws_horas[f"H{fila}"] = None

        ws_horas["G1"] = "OPERARIO (TOTALES)"
        ws_horas["H1"] = "TOTAL ACUMULADO"

        totales_acumulados = {}
        for r in range(2, ws_horas.max_row + 1):
            op_col_d = ws_horas[f"D{r}"].value
            hr_col_e = ws_horas[f"E{r}"].value
            if op_col_d and hr_col_e is not None:
                op_oficial = unificar_nombre(op_col_d)
                try:
                    horas_float = float(str(hr_col_e).replace(',', '.'))
                    totales_acumulados[op_oficial] = totales_acumulados.get(op_oficial, 0.0) + horas_float
                except ValueError:
                    pass

        fila_actual_totales = 2
        for operario, total_horas in sorted(totales_acumulados.items()):
            ws_horas[f"G{fila_actual_totales}"] = operario
            ws_horas[f"H{fila_actual_totales}"] = total_horas
            fila_actual_totales += 1

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # --- MOSTRAR BOTONES DE DESCARGA ---
        st.success("¡Todo procesado con éxito! Descarga tus archivos aquí abajo:")
        
        st.download_button(
            label="📥 Descargar Parte en PDF",
            data=pdf_output,
            file_name=f"Parte_{num_parte}_{fecha}.pdf",
            mime="application/pdf"
        )
        
        st.download_button(
            label="📥 Descargar Excel Actualizado",
            data=excel_buffer,
            file_name="Plantilla_Parte_Obra_Profesional.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Por favor, sube cualquier archivo Excel (.xlsx) en la sección de abajo para activar el botón de procesar.")
